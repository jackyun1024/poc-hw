#!/usr/bin/env python3
"""토스 포스 매출 엑셀 파일 검증 스크립트

Usage:
    python3 verify-sales-data.py file1.xlsx [file2.xlsx ...] [--password 0000]
"""
import sys
import os
import io
from datetime import datetime, timedelta

def open_xlsx(path, password="0000"):
    """암호화된 xlsx 열기"""
    import msoffcrypto
    import openpyxl

    with open(path, "rb") as f:
        # 암호화 여부 확인
        header = f.read(8)
        f.seek(0)

        if header[:4] == b'\xd0\xcf\x11\xe0':  # CDFV2 Encrypted
            file = msoffcrypto.OfficeFile(f)
            file.load_key(password=password)
            buf = io.BytesIO()
            file.decrypt(buf)
            buf.seek(0)
            return openpyxl.load_workbook(buf)
        else:
            return openpyxl.load_workbook(path)

def extract_dates(wb):
    """워크북에서 모든 날짜 추출"""
    dates = set()

    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row <= 2:
            continue

        # 날짜 컬럼 찾기
        date_col = None
        for c in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(1, c).value
            if val and any(k in str(val) for k in ['날짜', '일자', '기준일']):
                date_col = c
                break

        if date_col is None:
            continue

        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, date_col).value
            if isinstance(val, datetime):
                dates.add(val.date())
            elif isinstance(val, str):
                for fmt in ['%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d']:
                    try:
                        dates.add(datetime.strptime(val[:10], fmt).date())
                        break
                    except:
                        continue

    return dates

def main():
    files = []
    password = "0000"

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--password":
            i += 1
            password = args[i]
        else:
            files.append(args[i])
        i += 1

    if not files:
        print("Usage: verify-sales-data.py file1.xlsx [file2.xlsx ...] [--password 0000]")
        sys.exit(1)

    all_dates = set()

    for path in files:
        print(f"\n=== {os.path.basename(path)} ===")
        wb = open_xlsx(path, password)

        print(f"시트: {wb.sheetnames}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"  [{sname}] 행:{ws.max_row} 열:{ws.max_column}")

        dates = extract_dates(wb)
        if dates:
            sorted_dates = sorted(dates)
            print(f"날짜 범위: {sorted_dates[0]} ~ {sorted_dates[-1]}")
            print(f"고유 날짜: {len(dates)}일")
            all_dates |= dates
        else:
            print("데이터 없음")

    if all_dates:
        sorted_all = sorted(all_dates)
        print(f"\n{'='*50}")
        print(f"전체 범위: {sorted_all[0]} ~ {sorted_all[-1]}")
        print(f"고유 날짜: {len(all_dates)}일")

        expected = set()
        d = sorted_all[0]
        while d <= sorted_all[-1]:
            expected.add(d)
            d += timedelta(days=1)

        missing = sorted(expected - all_dates)
        print(f"예상 날짜: {len(expected)}일")

        if missing:
            print(f"\n⚠️  빠진 날짜 {len(missing)}개:")
            for m in missing:
                print(f"  {m} ({m.strftime('%a')})")
        else:
            print("\n✅ 빠진 날짜 없음!")

    print()

if __name__ == "__main__":
    main()
